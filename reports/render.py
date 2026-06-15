"""
render.py - turn report envelope(s) into management-ready Excel or PDF files.

Two modes share the same formatting helpers:
  * one report  -> one file        (render_excel / render_pdf)
  * many reports -> one "board pack" (render_excel_pack / render_pdf_pack):
    a single multi-sheet workbook, or a single PDF with a cover + a section per
    report - the artifact you hand to management.

Excel uses openpyxl (a project dependency); PDF uses reportlab. Both are optional
and degrade gracefully: a missing library is reported, never a crash.

Numbers are formatted for readers: financial figures get thousands separators,
years stay plain integers, periods keep their 3-decimal encoding, percent
columns show two decimals.
"""
from __future__ import annotations

import datetime as _dt
import os
import re

from . import safe_str as _safe


# --------------------------------------------------------------------------- #
# Shared number formatting
# --------------------------------------------------------------------------- #

# Arabic translations for report titles and descriptions
_REPORT_AR = {
    "Yearly P&L Summary": "ملخص الأرباح والخسائر السنوي",
    "Group-wide profit & loss by fiscal year (Actual).": "الأرباح والخسائر على مستوى المجموعة حسب السنة المالية (الفعلي).",
    "Regional P&L": "الأرباح والخسائر حسب المنطقة",
    "Profit & loss by region and year (Actual).": "الأرباح والخسائر حسب المنطقة والسنة (الفعلي).",
    "Product Group P&L": "الأرباح والخسائر حسب مجموعة المنتج",
    "Profit & loss by product group and year (Actual).": "الأرباح والخسائر حسب مجموعة المنتج والسنة (الفعلي).",
    "Country P&L": "الأرباح والخسائر حسب الدولة",
    "Profit & loss by country and year (Actual).": "الأرباح والخسائر حسب الدولة والسنة (الفعلي).",
    "Customer P&L": "الأرباح والخسائر حسب العميل",
    "Profit & loss by customer and year (Actual).": "الأرباح والخسائر حسب العميل والسنة (الفعلي).",
    "Year-over-Year Variance": "الانحراف السنوي",
    "Year-over-year P&L variance (Actual).": "الانحراف السنوي للأرباح والخسائر (الفعلي).",
    "Full-Year Outlook vs Prior Year": "توقعات السنة الكاملة مقابل السنة السابقة",
    "Full-year outlook benchmarked against prior year actual.": "توقعات السنة الكاملة مقارنة بالفعلي للسنة السابقة.",
    "Monthly Outlook Progression": "تطور التوقعات الشهرية",
    "Net sales and gross margin by month, flagged actual vs outlook.": "صافي المبيعات وهامش الربح شهريًا مع تمييز الفعلي مقابل التوقع.",
    "Import Validation Report": "تقرير التحقق من الاستيراد",
    "Row counts, null checks, duplicate grain, and source-coverage summary.": "عدد الصفوف، فحوص القيم الفارغة، تكرار المفتاح الرئيسي، وملخص تغطية المصدر.",
}

def _translate_report(title, description=""):
    """Translate report title and description to Arabic if available."""
    ar_title = _REPORT_AR.get(title, title)
    ar_desc = _REPORT_AR.get(description, description)
    return ar_title, ar_desc


def _is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _has_arabic(text):
    """True if the string contains any Arabic-script character (U+0600-06FF)."""
    return isinstance(text, str) and any("؀" <= ch <= "ۿ" for ch in text)


def envelope_has_arabic(envelope):
    """Detect Arabic anywhere in the headers or cells, to drive RTL layout."""
    if any(_has_arabic(col) for col in envelope.get("columns", [])):
        return True
    return any(_has_arabic(v) for row in envelope.get("rows", []) for v in row.values())


def excel_number_format(column):
    col = column.lower()
    if col == "year":
        return "0"
    if col == "period":
        return "0.000"
    if "pct" in col or "percent" in col:
        return "0.00"
    return "#,##0"


def pdf_format(column, value):
    if value is None or value == "":
        return ""
    if not _is_number(value):
        return str(value)
    col = column.lower()
    if col == "year":
        return str(int(value))
    if col == "period":
        return f"{value:.3f}"
    if "pct" in col or "percent" in col:
        return f"{value:.2f}"
    return f"{value:,.0f}"


def _safe_sheet_title(name):
    return re.sub(r"[\\/?*\[\]:]", "_", name)[:31] or "Report"


def _now():
    return _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def excel_available():
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:  # noqa: BLE001
        return False


def _excel_write_sheet(ws, envelope):
    """Write one report onto an existing worksheet (titled, formatted, frozen)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    columns = envelope["columns"]
    rows = envelope["rows"]
    ncols = max(len(columns), 1)

    # Arabic reports read right-to-left, so orient the whole sheet accordingly.
    if envelope_has_arabic(envelope):
        ws.sheet_view.rightToLeft = True

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title_cell = ws.cell(row=1, column=1, value=envelope["title"])
    title_cell.font = Font(size=14, bold=True)
    ws.cell(row=2, column=1, value=envelope.get("description", ""))
    src = envelope.get("source", {})
    ws.cell(row=3, column=1, value=(
        f"Generated {envelope.get('generated_at', '')} | "
        f"Source {src.get('database', '')} | {envelope.get('row_count', len(rows))} rows"))

    header_row = 5
    header_fill = PatternFill("solid", fgColor="1F3B57")
    header_font = Font(bold=True, color="FFFFFF")
    for c, column in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=c, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, start=header_row + 1):
        for c, column in enumerate(columns, start=1):
            value = row.get(column)
            cell = ws.cell(row=r, column=c, value=_safe(value) if not _is_number(value) else value)
            if _is_number(value):
                cell.number_format = excel_number_format(column)
                cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    for c, column in enumerate(columns, start=1):
        width = len(str(column))
        for row in rows:
            width = max(width, len(str(row.get(column, ""))))
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 10), 32)


def render_excel(envelope, out_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(envelope["report"])
    _excel_write_sheet(ws, envelope)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


def render_excel_pack(envelopes, out_path, title="Board Pack"):
    """One workbook: a Contents sheet + one sheet per report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    contents = wb.active
    contents.title = "Contents"
    contents.cell(row=1, column=1, value=title).font = Font(size=16, bold=True)
    contents.cell(row=2, column=1, value=f"Generated {_now()}")
    contents.cell(row=4, column=1, value="Report").font = Font(bold=True)
    contents.cell(row=4, column=2, value="Rows").font = Font(bold=True)
    for i, env in enumerate(envelopes, start=5):
        contents.cell(row=i, column=1, value=env["title"])
        contents.cell(row=i, column=2, value=env.get("row_count", len(env["rows"])))
    contents.column_dimensions["A"].width = 40

    for env in envelopes:
        ws = wb.create_sheet(_safe_sheet_title(env["report"]))
        _excel_write_sheet(ws, env)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------- #
# PDF
# --------------------------------------------------------------------------- #
_ARABIC_FONT_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "fonts", "NotoNaskhArabic.ttf",
)
_ARABIC_FONT_REGISTERED = False
_ARABIC_RESHAPER = None


def _arabic_shaper():
    """Lazy-load and return (reshape_fn, bidi_fn) or (None, None) if unavailable."""
    global _ARABIC_RESHAPER
    if _ARABIC_RESHAPER is not None:
        return _ARABIC_RESHAPER
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        config = arabic_reshaper.config_for_true_type_font(
            _ARABIC_FONT_PATH, arabic_reshaper.ENABLE_ALL_LIGATURES
        )
        reshaper = arabic_reshaper.ArabicReshaper(configuration=config)
        _ARABIC_RESHAPER = (reshaper.reshape, get_display)
        return _ARABIC_RESHAPER
    except Exception:
        _ARABIC_RESHAPER = (None, None)
        return _ARABIC_RESHAPER


def _shape_arabic(text):
    """Reshape + bidi an Arabic string for PDF rendering. Falls back to raw text."""
    reshape_fn, bidi_fn = _arabic_shaper()
    if not reshape_fn or not isinstance(text, str) or not _has_arabic(text):
        return text
    try:
        return bidi_fn(reshape_fn(text))
    except Exception:
        return text


def _register_arabic_font():
    """Register the Arabic font with ReportLab once."""
    global _ARABIC_FONT_REGISTERED
    if _ARABIC_FONT_REGISTERED:
        return
    if not os.path.exists(_ARABIC_FONT_PATH):
        return
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        pdfmetrics.registerFont(TTFont("Arabic", _ARABIC_FONT_PATH))
        _ARABIC_FONT_REGISTERED = True
    except Exception:
        pass
def pdf_available():
    try:
        import reportlab  # noqa: F401
        return True
    except Exception:  # noqa: BLE001
        return False


def _pdf_report_flowables(envelope, doc_width, is_arabic=False):
    """Reportlab flowables for one report: title, meta, table."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_RIGHT
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    columns = envelope["columns"]
    rows = envelope["rows"]
    src = envelope.get("source", {})

    if is_arabic:
        _register_arabic_font()
        body_font = "Arabic"
        ar_title_style = ParagraphStyle(
            'ArRptTitle', parent=styles['Title'],
            fontName='Arabic', alignment=TA_RIGHT, wordSpace='RTL'
        )
        ar_normal_style = ParagraphStyle(
            'ArRptNormal', parent=styles['Normal'],
            fontName='Arabic', alignment=TA_RIGHT
        )
        ar_title, ar_desc = _translate_report(envelope["title"], envelope.get("description", ""))
        title_text = _shape_arabic(ar_title)
        desc_text = _shape_arabic(ar_desc)
        meta_text = _shape_arabic(
            f"أُنشئ {envelope.get('generated_at', '')}  |  "
            f"المصدر {src.get('database', '')}  |  "
            f"{envelope.get('row_count', len(rows))} صف"
        )
        flow = [
            Paragraph(title_text, ar_title_style),
            Paragraph(desc_text, ar_normal_style),
            Paragraph(meta_text, ar_normal_style),
            Spacer(1, 6 * mm),
        ]
    else:
        body_font = "Helvetica"
        title_text = envelope["title"]
        desc_text = envelope.get("description", "")
        meta_text = (
            f"Generated {envelope.get('generated_at', '')} &nbsp;|&nbsp; "
            f"Source {src.get('database', '')} &nbsp;|&nbsp; "
            f"{envelope.get('row_count', len(rows))} rows"
        )
        flow = [
            Paragraph(title_text, styles["Title"]),
            Paragraph(desc_text, styles["Normal"]),
            Paragraph(meta_text, styles["Normal"]),
            Spacer(1, 6 * mm),
        ]

    font_size = 8 if len(columns) <= 8 else (7 if len(columns) <= 12 else 6)
    numeric_cols = set()
    for c, column in enumerate(columns):
        if rows and all(_is_number(row.get(column)) or row.get(column) in (None, "")
                        for row in rows):
            numeric_cols.add(c)

    if is_arabic:
        table_data = [[_shape_arabic(c) for c in columns]]
        table_data += [[pdf_format(col, row.get(col)) if _is_number(row.get(col))
                        else _shape_arabic(pdf_format(col, row.get(col)))
                        for col in columns] for row in rows]
    else:
        table_data = [columns] + [[pdf_format(col, row.get(col)) for col in columns]
                                  for row in rows]
    col_width = doc_width / max(len(columns), 1)
    table = Table(table_data, colWidths=[col_width] * len(columns), repeatRows=1)

    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3B57")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), body_font if is_arabic else "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), body_font),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for c in numeric_cols:
        style.append(("ALIGN", (c, 1), (c, -1), "RIGHT"))
    table.setStyle(TableStyle(style))
    flow.append(table)
    return flow


def _pdf_doc(out_path):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    return SimpleDocTemplate(out_path, pagesize=landscape(A4),
                             leftMargin=12 * mm, rightMargin=12 * mm,
                             topMargin=12 * mm, bottomMargin=12 * mm)


def render_pdf(envelope, out_path):
    """Render one report to PDF.

    Uses HTML/CSS -> WeasyPrint for Arabic content because ReportLab cannot
    shape connected Arabic glyphs. Falls back to ReportLab for English content.
    """
    is_arabic = envelope_has_arabic(envelope)
    try:
        weasyprint_available = False
        if is_arabic:
            from weasyprint import HTML  # noqa: F401
            weasyprint_available = True
    except Exception:
        weasyprint_available = False

    if is_arabic and weasyprint_available:
        return _render_pdf_single_html(envelope, out_path)

    doc = _pdf_doc(out_path)
    doc.build(_pdf_report_flowables(envelope, doc.width, is_arabic))
    return out_path


def _render_pdf_single_html(envelope, out_path):
    """High-quality single-report PDF using HTML/CSS via WeasyPrint."""
    from html import escape
    from pathlib import Path

    from weasyprint import HTML

    is_arabic = True
    dir_attr = "rtl"
    lang_attr = "ar"

    rpt_title, rpt_desc = _translate_report(envelope['title'], envelope.get('description', ''))
    src = envelope.get('source', {})
    meta = f"أُنشئت {envelope.get('generated_at', '')} · المصدر {src.get('database', '')} · {envelope.get('row_count', len(envelope['rows']))} صف"

    def _col_label(col):
        labels = {
            "year": "السنة", "version": "الإصدار", "period": "الفترة",
            "region_desc": "المنطقة", "product_group": "مجموعة المنتج",
            "country_desc": "الدولة", "customer_name": "العميل",
            "net_sales": "صافي المبيعات", "cogs": "تكلفة البضاعة المباعة",
            "gross_margin": "هامش الربح الإجمالي", "gross_margin_pct": "نسبة هامش الربح",
            "opex": "المصروفات التشغيلية", "operating_profit": "الربح التشغيلي",
            "operating_profit_pct": "نسبة الربح التشغيلي", "net_income": "صافي الدخل",
            "net_income_pct": "هامش صافي الدخل", "gross_sales": "إجمالي المبيعات",
            "returns": "المرتجعات", "sales_deduction": "الخصومات على المبيعات",
            "material_cost": "تكلفة المواد الخام", "sales_expense": "مصاريف البيع",
            "profit_before_tax": "الربح قبل الضرائب", "corporate_tax": "ضريبة الشركات",
            "royalty": "الروائب", "variance_pct": "نسبة التغير",
            "variance_abs": "قيمة التغير", "metric": "المقياس",
            "actual": "فعلي", "outlook": "التوقع", "combined": "مجموع", "scenario": "السيناريو",
            "description": "الوصف",
            "category": "الفئة", "item": "البند", "value": "القيمة", "status": "الحالة",
        }
        return labels.get(col, col)

    def _cell(row, col):
        val = row.get(col)
        if val is None:
            return ""
        if isinstance(val, (int, float)):
            return f'{val:,.0f}'
        return escape(str(val))

    css = """
    @page { size: A4 landscape; margin: 14mm; }
    body {
      font-family: "Noto Naskh Arabic", "Noto Sans Arabic", serif;
      color: #172033; background: white; font-size: 10pt; line-height: 1.55;
      margin: 0; padding: 0;
    }
    .rtl { direction: rtl; text-align: right; }
    h1 { font-family: "Noto Sans Arabic", sans-serif; font-size: 30px; margin: 0 0 6px; color: #12395a; }
    h2 { font-family: "Noto Sans Arabic", sans-serif; font-size: 18px; margin: 16px 0 6px; color: #12395a; }
    .meta { color: #667085; margin-bottom: 14px; }
    .intro { margin: 10px 0 16px; }
    table { width: 100%; border-collapse: collapse; margin-top: 10px; direction: rtl; }
    th { background: #12395a; color: white; font-family: "Noto Sans Arabic", sans-serif; padding: 7px 9px; text-align: right; }
    td { padding: 6px 9px; border: 1px solid #d9e2ec; text-align: right; }
    tr:nth-child(even) td { background: #f7f9fc; }
    .num { direction: ltr; unicode-bidi: isolate; text-align: left; font-family: "Noto Sans Arabic", sans-serif; }
    """

    parts = [
        f'<!doctype html><html lang="{lang_attr}" dir="{dir_attr}"><head><meta charset="utf-8"><style>{css}</style></head><body>',
        f'<div class="rtl"><h1>{escape(rpt_title)}</h1>',
        f'<div class="intro">{escape(rpt_desc)}</div>',
        f'<div class="meta">{escape(meta)}</div>',
        '<table><thead><tr>',
    ]
    for col in envelope['columns']:
        parts.append(f'<th>{escape(_col_label(col))}</th>')
    parts.append('</tr></thead><tbody>')
    for row in envelope['rows']:
        parts.append('<tr>')
        for col in envelope['columns']:
            cls = 'num' if isinstance(row.get(col), (int, float)) else ''
            parts.append(f'<td class="{cls}">{_cell(row, col)}</td>')
        parts.append('</tr>')
    parts.append('</tbody></table></div></body></html>')

    html = "\n".join(parts)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    HTML(string=html, base_url=str(Path(out_path).parent)).write_pdf(out_path)
    return out_path


def render_pdf_pack(envelopes, out_path, title="Board Pack"):
    """One PDF: a cover page + a section per report, page-broken between.

    For Arabic/RTL titles, we use the HTML/CSS -> WeasyPrint path because
    ReportLab cannot render connected Arabic glyphs reliably.  For English
    titles we keep the original ReportLab path.
    """
    is_arabic = _has_arabic(title)
    try:
        weasyprint_available = False
        if is_arabic:
            from weasyprint import HTML  # noqa: F401
            weasyprint_available = True
    except Exception:
        weasyprint_available = False

    if is_arabic and weasyprint_available:
        return render_pdf_pack_html(envelopes, out_path, title=title)

    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.enums import TA_RIGHT
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, Spacer

    styles = getSampleStyleSheet()
    doc = _pdf_doc(out_path)

    if is_arabic:
        _register_arabic_font()
        # Build dedicated Arabic paragraph styles (RTL)
        ar_title_style = ParagraphStyle(
            'ArabicTitle', parent=styles['Title'],
            fontName='Arabic', alignment=TA_RIGHT, wordSpace='RTL'
        )
        ar_heading_style = ParagraphStyle(
            'ArabicHeading', parent=styles['Heading2'],
            fontName='Arabic', alignment=TA_RIGHT
        )
        ar_normal_style = ParagraphStyle(
            'ArabicNormal', parent=styles['Normal'],
            fontName='Arabic', alignment=TA_RIGHT
        )
        cover_title = _shape_arabic(title)
        cover_subtitle = _shape_arabic(f"أُنشئ في {_now()}")
        cover_contents = _shape_arabic("المحتويات")
        story = [
            Paragraph(cover_title, ar_title_style),
            Paragraph(cover_subtitle, ar_normal_style),
            Spacer(1, 8 * mm),
            Paragraph(cover_contents, ar_heading_style),
        ]
    else:
        story = [
            Paragraph(title, styles["Title"]),
            Paragraph(f"Generated {_now()}", styles["Normal"]),
            Spacer(1, 8 * mm),
            Paragraph("Contents", styles["Heading2"]),
        ]

    for env in envelopes:
        if is_arabic:
            ar_env_title, _ = _translate_report(env['title'])
            env_text = _shape_arabic(f"• {ar_env_title} ({env.get('row_count', len(env['rows']))})")
            story.append(Paragraph(env_text, ar_normal_style))
        else:
            story.append(Paragraph(
                f"&bull; {env['title']} ({env.get('row_count', len(env['rows']))} rows)",
                styles["Normal"]))

    for env in envelopes:
        story.append(PageBreak())
        story.extend(_pdf_report_flowables(env, doc.width, is_arabic))

    doc.build(story)
    return out_path


def render_pdf_pack_html(envelopes, out_path, title="Board Pack"):
    """High-quality Arabic/RTL PDF board pack using HTML/CSS via WeasyPrint.

    Falls back to the ReportLab renderer if WeasyPrint is unavailable.
    """
    from html import escape
    from pathlib import Path

    try:
        from weasyprint import HTML
    except Exception:
        return render_pdf_pack(envelopes, out_path, title)

    is_arabic = _has_arabic(title)
    if is_arabic:
        title, _ = _translate_report(title)
        dir_attr = "rtl"
        lang_attr = "ar"
    else:
        dir_attr = "ltr"
        lang_attr = "en"

    def _col_label(col):
        labels = {
            "year": "السنة",
            "version": "الإصدار",
            "period": "الفترة",
            "region_desc": "المنطقة",
            "product_group": "مجموعة المنتج",
            "country_desc": "الدولة",
            "customer_name": "العميل",
            "net_sales": "صافي المبيعات",
            "cogs": "تكلفة البضاعة المباعة",
            "gross_margin": "هامش الربح الإجمالي",
            "gross_margin_pct": "نسبة هامش الربح",
            "opex": "المصروفات التشغيلية",
            "operating_profit": "الربح التشغيلي",
            "operating_profit_pct": "نسبة الربح التشغيلي",
            "net_income": "صافي الدخل",
            "net_income_pct": "هامش صافي الدخل",
            "gross_sales": "إجمالي المبيعات",
            "returns": "المرتجعات",
            "sales_deduction": "الخصومات على المبيعات",
            "material_cost": "تكلفة المواد الخام",
            "sales_expense": "مصاريف البيع",
            "profit_before_tax": "الربح قبل الضرائب",
            "corporate_tax": "ضريبة الشركات",
            "royalty": "الروائب",
            "variance_pct": "نسبة التغير",
            "variance_abs": "قيمة التغير",
            "metric": "المقياس",
            "actual": "فعلي",
            "outlook": "التوقع",
            "combined": "مجموع",
            "scenario": "السيناريو",
            "description": "الوصف",
            "category": "الفئة", "item": "البند", "value": "القيمة", "status": "الحالة",
        }
        return labels.get(col, col)

    def _cell(row, col):
        val = row.get(col)
        if val is None:
            return ""
        if isinstance(val, (int, float)):
            return f'{val:,.0f}'
        return escape(str(val))

    css = """
    @page { size: A4 landscape; margin: 14mm; }
    body {
      font-family: "Noto Naskh Arabic", "Noto Sans Arabic", serif;
      color: #172033; background: white; font-size: 10pt; line-height: 1.55;
      margin: 0; padding: 0;
    }
    .rtl { direction: rtl; text-align: right; }
    .ltr { direction: ltr; text-align: left; }
    .cover { border: 1.5px solid #d9e2ec; border-radius: 12px; padding: 22px 28px; background: #fbfcff; }
    h1 { font-family: "Noto Sans Arabic", sans-serif; font-size: 32px; margin: 0 0 6px; color: #12395a; }
    h2 { font-family: "Noto Sans Arabic", sans-serif; font-size: 20px; margin: 18px 0 8px; color: #12395a; page-break-after: avoid; }
    .meta { color: #667085; margin-bottom: 14px; }
    .section { margin-top: 10px; page-break-before: always; }
    .section:first-of-type { page-break-before: auto; }
    table { width: 100%; border-collapse: collapse; margin-top: 10px; direction: rtl; }
    th { background: #12395a; color: white; font-family: "Noto Sans Arabic", sans-serif; padding: 7px 9px; text-align: right; }
    td { padding: 6px 9px; border: 1px solid #d9e2ec; text-align: right; }
    tr:nth-child(even) td { background: #f7f9fc; }
    .num { direction: ltr; unicode-bidi: isolate; text-align: left; font-family: "Noto Sans Arabic", sans-serif; }
    .toc-item { margin: 5px 0; }
    .intro { margin: 10px 0 16px; }
    .status-ok { color: #0f6e31; font-weight: bold; }
    .status-warn { color: #b45309; font-weight: bold; }
    """

    parts = []
    parts.append(f'<!doctype html><html lang="{lang_attr}" dir="{dir_attr}"><head><meta charset="utf-8"><style>{css}</style></head><body>')

    # Cover
    if is_arabic:
        parts.append(f'<div class="cover rtl"><h1>{escape(title)}</h1><div class="meta">أُنشئت في {_now()}</div></div>')
        parts.append(f'<div class="rtl" style="margin-top:18px;"><h2>محتويات الحزمة</h2>')
    else:
        parts.append(f'<div class="cover ltr"><h1>{escape(title)}</h1><div class="meta">Generated {_now()}</div></div>')
        parts.append(f'<div class="ltr" style="margin-top:18px;"><h2>Contents</h2>')

    for env in envelopes:
        if is_arabic:
            rpt_title, _ = _translate_report(env['title'])
            parts.append(f'<div class="toc-item">{escape(rpt_title)} ({env.get("row_count", len(env["rows"]))})</div>')
        else:
            parts.append(f'<div class="toc-item">{escape(env["title"])} ({env.get("row_count", len(env["rows"]))} rows)</div>')
    parts.append('</div>')

    # Reports
    for env in envelopes:
        if is_arabic:
            rpt_title, rpt_desc = _translate_report(env['title'], env.get('description', ''))
            src = env.get('source', {})
            meta = f"أُنشئت {env.get('generated_at', '')} · المصدر {src.get('database', '')} · {env.get('row_count', len(env['rows']))} صف"
        else:
            rpt_title = env['title']
            rpt_desc = env.get('description', '')
            src = env.get('source', {})
            meta = f"Generated {env.get('generated_at', '')} · Source {src.get('database', '')} · {env.get('row_count', len(env['rows']))} rows"

        parts.append(f'<div class="section rtl"><h2>{escape(rpt_title)}</h2>')
        parts.append(f'<div class="intro">{escape(rpt_desc)}</div>')
        parts.append(f'<div class="meta">{escape(meta)}</div>')
        parts.append('<table><thead><tr>')
        for col in env['columns']:
            parts.append(f'<th>{escape(_col_label(col))}</th>')
        parts.append('</tr></thead><tbody>')
        for row in env['rows']:
            parts.append('<tr>')
            for col in env['columns']:
                cls = 'num' if isinstance(row.get(col), (int, float)) else ''
                parts.append(f'<td class="{cls}">{_cell(row, col)}</td>')
            parts.append('</tr>')
        parts.append('</tbody></table></div>')

    # Source confidence / data lineage page
    if is_arabic:
        validation_env = next((e for e in envelopes if e.get('report') == 'import_validation'), None)
        if validation_env:
            summary_rows = [r for r in validation_env['rows'] if r['category'] == 'Summary']
            parts.append('<div class="section rtl"><h2>ثقة المصدر وملخص التحقق</h2>')
            parts.append('<div class="intro">هذا الملخص مؤسس على البيانات المستخرجة من قاعدة البيانات وحسابات الجودة الأساسية.</div>')
            parts.append('<table><thead><tr><th>البند</th><th>القيمة</th><th>الحالة</th></tr></thead><tbody>')
            for row in summary_rows:
                status_class = 'status-ok' if row['status'] == 'OK' else 'status-warn'
                parts.append(f'<tr><td>{escape(row["item"])}</td><td class="num">{row["value"]}</td><td class="{status_class}">{escape(row["status"])}</td></tr>')
            parts.append('</tbody></table></div>')

    parts.append('</body></html>')

    html = "\n".join(parts)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    HTML(string=html, base_url=str(Path(out_path).parent)).write_pdf(out_path)
    return out_path

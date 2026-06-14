"""
Tests for the Excel/PDF renderers (no pytest required).

Excel is always tested (openpyxl is a project dependency). PDF is tested only
when reportlab is importable, and skipped cleanly otherwise.

Run:  python3 -m reports.test_render
"""
from __future__ import annotations

import os
import tempfile

from . import render

ENVELOPE = {
    "report": "yearly_pl",
    "title": "Yearly P&L Summary",
    "description": "Group-wide P&L by fiscal year.",
    "generated_at": "2026-06-14T09:00:00Z",
    "source": {"database": "pl_detail.db", "rows_in_ledger": 7560},
    "columns": ["year", "net_sales", "gross_margin", "net_sales_pct"],
    "row_count": 2,
    "rows": [
        {"year": 2024, "net_sales": 114248976.0, "gross_margin": 40000000.0, "net_sales_pct": 6.64},
        {"year": 2025, "net_sales": 120251315.0, "gross_margin": 42000000.0, "net_sales_pct": 5.25},
    ],
}


def test_number_formatting():
    assert render.pdf_format("net_sales", 114248976.0) == "114,248,976"
    assert render.pdf_format("year", 2025) == "2025"
    assert render.pdf_format("period", 2025.003) == "2025.003"
    assert render.pdf_format("net_sales_pct", 6.64) == "6.64"
    assert render.pdf_format("region_desc", "Africa") == "Africa"
    assert render.pdf_format("net_sales", None) == ""
    assert render.excel_number_format("net_sales") == "#,##0"
    assert render.excel_number_format("year") == "0"


def test_excel():
    assert render.excel_available(), "openpyxl should be installed"
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "yearly_pl.xlsx")
        render.render_excel(ENVELOPE, path)
        assert os.path.getsize(path) > 0

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws["A1"].value == "Yearly P&L Summary"
        # Header row is row 5; data begins row 6 with the real numeric value.
        assert ws.cell(row=5, column=1).value == "year"
        assert ws.cell(row=6, column=1).value == 2024
        assert ws.cell(row=6, column=2).value == 114248976.0
        assert ws.cell(row=6, column=2).number_format == "#,##0"


def test_pdf():
    if not render.pdf_available():
        print("SKIP: reportlab not installed; PDF render not tested here.")
        return
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "yearly_pl.pdf")
        render.render_pdf(ENVELOPE, path)
        assert os.path.getsize(path) > 0
        with open(path, "rb") as fh:
            assert fh.read(5) == b"%PDF-", "not a valid PDF file"


SECOND = {
    "report": "regional_pl", "title": "Regional P&L", "description": "By region.",
    "generated_at": "2026-06-14T09:00:00Z",
    "source": {"database": "pl_detail.db", "rows_in_ledger": 7560},
    "columns": ["year", "region_desc", "net_sales"], "row_count": 1,
    "rows": [{"year": 2025, "region_desc": "Africa", "net_sales": 1200.0}],
}


def test_excel_pack():
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "board-pack.xlsx")
        render.render_excel_pack([ENVELOPE, SECOND], path, title="Test Pack")
        import openpyxl
        wb = openpyxl.load_workbook(path)
        # Contents sheet + one per report.
        assert wb.sheetnames == ["Contents", "yearly_pl", "regional_pl"], wb.sheetnames
        assert wb["Contents"]["A1"].value == "Test Pack"
        assert wb["regional_pl"].cell(row=6, column=3).value == 1200.0


def test_pdf_pack():
    if not render.pdf_available():
        print("SKIP: reportlab not installed; PDF pack not tested here.")
        return
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "board-pack.pdf")
        render.render_pdf_pack([ENVELOPE, SECOND], path, title="Test Pack")
        with open(path, "rb") as fh:
            assert fh.read(5) == b"%PDF-"
        assert os.path.getsize(path) > 0


def main():
    test_number_formatting()
    test_excel()
    test_pdf()
    test_excel_pack()
    test_pdf_pack()
    print("render tests passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
